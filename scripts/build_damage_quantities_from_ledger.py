#!/usr/bin/env python3
"""재해대장보고서 엑셀 → damage_recovery_events_seed.json 피해·복구 정량치 보강 전처리.

용도
  `data/seed/damage_recovery_events_seed.json`의 ACTUAL_BACKED 레코드 중
  `damage.quantities_status == "NOT_AVAILABLE"` 인 건에 대하여,
  행정안전부 재해대장보고서(2010~2025) 원시 엑셀의 **공공시설 피해금액·복구비**를
  `재난년도 + 시도·시군구 + 재난명(사상·월)` 기준으로 집계해 채운다.

원칙(보수적 매칭)
  - 연도·지역·재난유형이 모두 명확히 일치하고 후보가 유일할 때만 매칭한다.
  - 후보가 0건이거나 2건 이상이면 매칭하지 않고 `NOT_AVAILABLE`을 유지한다.
  - 집계 범위는 **시군구 전체**이며 위험지구(district) 단위 금액이 아니다.
  - 재해대장은 **공공시설** 피해·복구 대장이므로 인명·사유재산 피해는 포함하지 않는다.
  - 산출값은 **과거 확정/취합 집계값**이며 피해예측이 아니다(`is_prediction=false`).
  - 금액 단위는 원시 엑셀 그대로 **천원(KRW_THOUSAND)** 이다.

Seed 계약 보호
  - record_id·event_id·evidence_id 등 기존 ID와 필드명을 변경하지 않는다.
  - 기존 `damage.description`, `damage.private_facility`, `damage.public_facility`,
    `damage.agriculture` 는 손대지 않는다(유사도 damage_pattern 요인 입력값이므로
    수정 시 랭킹 회귀가 발생한다).
  - 레코드 최상위 `data_status`/`official_data`/`provisional`은 유지한다
    (재해대장 유래 실집계임은 `damage.ledger_aggregate`와 evidence 항목에 표기).

사용
  python scripts/build_damage_quantities_from_ledger.py [--ledger <xlsx>] [--check] [--report <path>]

  --check : Seed를 쓰지 않고 매칭 결과만 출력한다.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LEDGER = ROOT / '메타데이터 참고자료(T3Q)' / '20260708_2010~2025재해대장보고서_New.xlsx'
SEED_PATH = ROOT / 'data/seed/damage_recovery_events_seed.json'
PUBLIC_SEED_PATH = ROOT / 'apps/web/public/seed/damage_recovery_events_seed.json'

SOURCE_DOCUMENT = '행정안전부 재해대장보고서(2010~2025)'
SOURCE_FILE = '20260708_2010~2025재해대장보고서_New.xlsx'
SOURCE_DATASET_ID = 'DISASTER_LEDGER_REPORT_2010_2025'
GENERATED_BY = 'scripts/build_damage_quantities_from_ledger.py'
CURRENCY_UNIT = 'KRW_THOUSAND'
UNIT_LABEL = '천원'
QUANTITIES_STATUS = 'AVAILABLE_PUBLIC_FACILITY_LEDGER'
LEDGER_BADGE = '재해대장 공공시설 피해·복구 집계 반영'

# 엑셀 2행 헤더 기준 0-based 컬럼 위치
COL = {
    'year': 0, 'disaster_name': 1, 'sido': 2, 'sigungu': 3, 'dong': 4,
    'location': 5, 'ministry': 6, 'facility_type': 7, 'facility_grade': 8,
    'facility_name': 9, 'draft_stage': 10, 'recovery_stage': 11,
    'recovery_type': 12,
    'damage_local': 13, 'recovery_local': 14,
    'damage_central': 15, 'recovery_central': 16,
    'registration': 17,
}

# 대상 시군구(행정구역코드 → 엑셀 표기). 시도 표기는 개편 전후 명칭을 모두 허용한다.
REGIONS = {
    '41430': {'sigungu': '의왕시', 'sido': ('경기도',)},
    '47190': {'sigungu': '구미시', 'sido': ('경상북도',)},
    '45190': {'sigungu': '남원시', 'sido': ('전라북도', '전북특별자치도')},
}

# 재난명 → 재난유형 판정 키워드 (Seed disaster_type과 대조)
NAME_TYPE_KEYWORDS = (
    ('TYPH', ('태풍',)),
    ('FLOOD', ('호우', '홍수', '집중호우', '대우', '폭우')),
    ('SNOW', ('대설',)),
    ('WIND', ('강풍', '풍랑')),
    ('EQ', ('지진',)),
)
MONTH_TOKEN = re.compile(r'(?<!\d)(\d{1,2})\.\s*(\d{1,2})')
TYPHOON_NAME = re.compile(r'태풍\s*([가-힣]{2,6})')


def read_ledger(path: Path):
    """엑셀 1시트를 (헤더 2행 제외) 그대로 dict 목록으로 읽는다."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - 실행환경 안내용
        raise SystemExit('openpyxl이 필요합니다: pip install openpyxl')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for raw in ws.iter_rows(min_row=3, values_only=True):
        if raw[COL['year']] is None:
            continue
        rows.append(raw)
    wb.close()
    return rows


def num(value) -> int:
    return int(value) if isinstance(value, (int, float)) else 0


def name_types(name: str) -> set[str]:
    return {code for code, words in NAME_TYPE_KEYWORDS if any(w in name for w in words)}


def name_months(name: str) -> set[int]:
    months = set()
    for m, _d in MONTH_TOKEN.findall(name):
        value = int(m)
        if 1 <= value <= 12:
            months.add(value)
    return months


def group_ledger(rows):
    """(시군구, 재난년도, 재난명) 단위 집계."""
    groups: dict[tuple[str, str, str], dict] = {}
    for raw in rows:
        sgg = raw[COL['sigungu']]
        sido = raw[COL['sido']]
        target = next((code for code, meta in REGIONS.items()
                       if meta['sigungu'] == sgg and sido in meta['sido']), None)
        if target is None:
            continue
        key = (target, str(raw[COL['year']]).strip(), str(raw[COL['disaster_name']]).strip())
        g = groups.setdefault(key, {
            'admin_code': target, 'sido': sido, 'sigungu': sgg,
            'row_count': 0,
            'damage_local': 0, 'recovery_local': 0,
            'damage_central': 0, 'recovery_central': 0,
            'facility': defaultdict(lambda: {
                'row_count': 0, 'damage_local': 0, 'recovery_local': 0,
                'damage_central': 0, 'recovery_central': 0,
            }),
            'recovery_stage': Counter(), 'registration': Counter(),
            'ministry': Counter(), 'dong': Counter(),
        })
        g['row_count'] += 1
        f = g['facility'][str(raw[COL['facility_type']] or '미분류')]
        for field in ('damage_local', 'recovery_local', 'damage_central', 'recovery_central'):
            value = num(raw[COL[field]])
            g[field] += value
            f[field] += value
        f['row_count'] += 1
        g['recovery_stage'][str(raw[COL['recovery_stage']] or '미기재')] += 1
        g['registration'][str(raw[COL['registration']] or '미기재')] += 1
        g['ministry'][str(raw[COL['ministry']] or '미기재')] += 1
        g['dong'][str(raw[COL['dong']] or '미기재')] += 1
    return groups


def match_record(record: dict, groups: dict) -> tuple[tuple | None, str]:
    """Seed 레코드 1건에 대응하는 재해대장 그룹을 보수적으로 찾는다."""
    admin_code = str(record.get('admin_code') or '')
    if admin_code not in REGIONS:
        return None, 'NO_MATCH: 대상 시군구가 아님'
    occurred = str(record.get('occurred_from') or '')
    if len(occurred) < 7:
        return None, 'NO_MATCH: 발생시각 미확보'
    year, month = occurred[:4], int(occurred[5:7])
    seed_type = str(record.get('disaster_type') or '')
    event_name = str(record.get('event_name') or '')

    candidates = [k for k in groups if k[0] == admin_code and k[1] == year]
    if not candidates:
        return None, f'NO_MATCH: {year}년 {REGIONS[admin_code]["sigungu"]} 재해대장 자료 없음'

    kept = []
    for key in candidates:
        ledger_name = key[2]
        types = name_types(ledger_name)
        if seed_type and types and seed_type not in types:
            continue
        months = name_months(ledger_name)
        if months:
            # 재난명에 월이 표기된 경우 Seed 발생월이 표기월 범위 안에 있어야 한다.
            if not (min(months) <= month <= max(months)):
                continue
        elif seed_type == 'TYPH':
            # 월 표기가 없는 태풍 사상은 태풍 고유명이 Seed 사건명과 일치해야 한다.
            typhoon = TYPHOON_NAME.search(ledger_name)
            if not typhoon or typhoon.group(1) not in event_name:
                continue
        else:
            continue
        kept.append(key)

    if len(kept) == 1:
        return kept[0], 'MATCH: 재난년도·시군구·재난유형·재난명(월/사상) 일치, 후보 유일'
    if not kept:
        return None, f'NO_MATCH: {year}년 후보 {len(candidates)}건 중 재난유형·시기 일치 없음'
    names = ', '.join(k[2] for k in kept)
    return None, f'NO_MATCH: 후보 다중({len(kept)}건: {names}) — 보수적으로 미매칭 유지'


def krw(value: int) -> str:
    return f'{value:,}{UNIT_LABEL}'


def eok(value: int) -> str:
    """천원 단위 값을 억원 환산 문구로 보조 표기한다(1억원 = 100,000천원)."""
    return f'약 {value / 100000:,.1f}억원'


def confirmation_status(group) -> str:
    """복구단계 구성으로 '중앙확정' 여부를 판정한다. 확정 전 단계를 확정으로 표기하지 않는다."""
    stages = set(group['recovery_stage'])
    return '중앙확정' if stages == {'중앙확정'} else '중앙확정 이전 단계 포함'


def build_aggregate(key, group, today: str) -> dict:
    facility = sorted(
        ({'facility_type': name, **{k: v for k, v in data.items()}}
         for name, data in group['facility'].items()),
        key=lambda row: (-row['damage_central'], -row['recovery_central'], row['facility_type']),
    )
    return {
        'source_document': SOURCE_DOCUMENT,
        'source_dataset_id': SOURCE_DATASET_ID,
        'source_file': SOURCE_FILE,
        'source_organization': '행정안전부',
        'official_data': True,
        'data_status': 'actual',
        'value_status': 'actual',
        'is_prediction': False,
        'reference_only': True,
        'record_scope': '공공시설 재해대장(피해시설 단위) 집계 — 인명피해·사유재산 피해는 포함하지 않음',
        'aggregation_scope': '시군구 전체 합계(위험지구·행정동 단위 귀속 금액 아님)',
        'aggregation_basis': '재난년도 + 시도·시군구 + 재난명 일치 행 전체 합계',
        'match_rule': '재난년도 + 시도·시군구 + 재난유형 + 재난명(월/태풍 사상) 일치 후보가 유일한 경우에만 매칭',
        'matched_disaster_year': key[1],
        'matched_disaster_name': key[2],
        'matched_sido': group['sido'],
        'matched_sigungu': group['sigungu'],
        'matched_row_count': group['row_count'],
        'currency_unit': CURRENCY_UNIT,
        'unit_label': UNIT_LABEL,
        'damage_amount_local_report': group['damage_local'],
        'recovery_cost_local_report': group['recovery_local'],
        'damage_amount_central_confirmed': group['damage_central'],
        'recovery_cost_central_confirmed': group['recovery_central'],
        'confirmation_status': confirmation_status(group),
        'recovery_stage_counts': dict(group['recovery_stage'].most_common()),
        'registration_counts': dict(group['registration'].most_common()),
        'affected_dong_count': len(group['dong']),
        'facility_breakdown': facility,
        'generated_by': GENERATED_BY,
        'generated_at': today,
    }


def build_note(aggregate: dict) -> str:
    dc = aggregate['damage_amount_central_confirmed']
    rc = aggregate['recovery_cost_central_confirmed']
    dl = aggregate['damage_amount_local_report']
    rl = aggregate['recovery_cost_local_report']
    stage = ' / '.join(f'{k} {v}건' for k, v in aggregate['recovery_stage_counts'].items())
    local_note = ''
    if dl != dc or rl != rc:
        local_note = f' 시군구 보고 기준은 피해 {krw(dl)}·복구비 {krw(rl)}.'
    confirmed = aggregate['confirmation_status'] == '중앙확정'
    closing = ('과거 확정 집계값이며 피해예측이 아니다.' if confirmed
               else '중앙확정 이전 단계(취합중)를 포함한 과거 집계값이며 피해예측이 아니다.')
    return (
        f'{SOURCE_DOCUMENT} 집계 — {aggregate["matched_sigungu"]} '
        f'{aggregate["matched_disaster_year"]}년 「{aggregate["matched_disaster_name"]}」 '
        f'재해대장 {aggregate["matched_row_count"]}건, '
        f'공공시설 피해금액 {krw(dc)}({eok(dc)}) · 복구비 {krw(rc)}({eok(rc)}) '
        f'[중앙보고 기준, {stage}].{local_note} '
        f'금액단위 {UNIT_LABEL}. 시군구 전체 합계이며 위험지구 단위 금액이 아니고, '
        f'인명·사유재산 피해는 미포함이다. {closing}'
    )


def build_evidence(record: dict, aggregate: dict) -> dict:
    dc = aggregate['damage_amount_central_confirmed']
    rc = aggregate['recovery_cost_central_confirmed']
    return {
        'evidence_id': f'EVD-{record["event_id"]}-LEDGER-01',
        'source_type': 'DISASTER_LEDGER_REPORT',
        'data_status': 'actual',
        'title': f'{SOURCE_DOCUMENT} 공공시설 피해·복구 집계',
        'page': None,
        'passage_id': None,
        'excerpt': (
            f'{aggregate["matched_sido"]} {aggregate["matched_sigungu"]} '
            f'{aggregate["matched_disaster_year"]}년 「{aggregate["matched_disaster_name"]}」 '
            f'재해대장 {aggregate["matched_row_count"]}건 — 공공시설 피해금액 {krw(dc)} · '
            f'복구비 {krw(rc)}(시군구 전체 합계, 단위 {UNIT_LABEL}, '
            f'복구단계 {aggregate["confirmation_status"]}, 과거 집계값·피해예측 아님)'
        ),
        'lineage': {
            'source_file': SOURCE_FILE,
            'source_organization': '행정안전부',
            'dataset_id': SOURCE_DATASET_ID,
            'disaster_year': aggregate['matched_disaster_year'],
            'disaster_name': aggregate['matched_disaster_name'],
            'sido': aggregate['matched_sido'],
            'sigungu': aggregate['matched_sigungu'],
            'row_count': aggregate['matched_row_count'],
            'currency_unit': CURRENCY_UNIT,
            'aggregation_scope': aggregate['aggregation_scope'],
            'official_data': True,
            'is_prediction': False,
            'generated_by': GENERATED_BY,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--ledger', default=str(DEFAULT_LEDGER), help='재해대장보고서 엑셀 경로')
    parser.add_argument('--check', action='store_true', help='Seed를 쓰지 않고 매칭 결과만 출력')
    parser.add_argument('--report', default=None, help='매칭 결과 JSON 저장 경로(선택)')
    args = parser.parse_args()

    ledger_path = Path(args.ledger)
    if not ledger_path.exists():
        print(f'FAIL 재해대장 엑셀을 찾을 수 없습니다: {ledger_path}')
        return 2

    today = date.today().isoformat()
    groups = group_ledger(read_ledger(ledger_path))
    seed = json.loads(SEED_PATH.read_text(encoding='utf-8'))

    matched, skipped = [], []
    for record in seed['records']:
        if record.get('data_status') != 'actual_backed':
            skipped.append({'record_id': record['record_id'], 'reason': 'SKIP: SYNTHETIC_DEMO 레코드'})
            continue
        key, reason = match_record(record, groups)
        if key is None:
            skipped.append({'record_id': record['record_id'], 'event_id': record['event_id'], 'reason': reason})
            continue
        aggregate = build_aggregate(key, groups[key], today)
        matched.append({
            'record_id': record['record_id'], 'event_id': record['event_id'], 'reason': reason,
            'disaster_year': key[1], 'disaster_name': key[2], 'row_count': groups[key]['row_count'],
            'damage_amount_central_confirmed': aggregate['damage_amount_central_confirmed'],
            'recovery_cost_central_confirmed': aggregate['recovery_cost_central_confirmed'],
        })
        if args.check:
            continue

        damage = record['damage']
        damage['quantities_status'] = QUANTITIES_STATUS
        damage['damage_note'] = build_note(aggregate)
        damage['ledger_aggregate'] = aggregate
        damage['is_prediction'] = False

        evidence_item = build_evidence(record, aggregate)
        evidence = [e for e in record.get('evidence', []) if e.get('evidence_id') != evidence_item['evidence_id']]
        evidence.append(evidence_item)
        record['evidence'] = evidence

        badges = [b for b in record.get('display_badges', []) if b != LEDGER_BADGE]
        badges.insert(1, LEDGER_BADGE)
        record['display_badges'] = badges

    if not args.check:
        seed['as_of'] = today
        seed['notice'] = (
            '9건은 자연재해저감종합계획의 피해이력 문구를 구조화한 ACTUAL_BACKED 레코드이며, '
            '6건은 유사도·대응비교·API 교체계약 검증용 SYNTHETIC_DEMO이다. '
            f'ACTUAL_BACKED {len(matched)}건의 damage.ledger_aggregate는 {SOURCE_DOCUMENT}에서 '
            '재난년도·시군구·재난명 기준으로 집계한 공공시설 피해금액·복구비 실집계값(단위 천원, 시군구 전체 합계)이며 '
            '과거 집계이지 피해예측이 아니다. 나머지 항목은 실제 NDMS 데이터 또는 피해예측 결과가 아니다.'
        )
        text = json.dumps(seed, ensure_ascii=False, indent=2) + '\n'
        SEED_PATH.write_text(text, encoding='utf-8')
        if PUBLIC_SEED_PATH.exists():
            PUBLIC_SEED_PATH.write_text(text, encoding='utf-8')

    report = {
        'generated_at': today,
        'generated_by': GENERATED_BY,
        'ledger_file': SOURCE_FILE,
        'currency_unit': CURRENCY_UNIT,
        'matched_count': len(matched),
        'skipped_count': len(skipped),
        'matched': matched,
        'skipped': skipped,
    }
    if args.report:
        Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

    mode = 'CHECK' if args.check else 'WRITE'
    print(f'{mode} 재해대장 정량치 반영: 매칭 {len(matched)}건 / 미매칭·제외 {len(skipped)}건')
    for row in matched:
        print(f'  [MATCH] {row["record_id"]} ← {row["disaster_year"]} {row["disaster_name"]} '
              f'({row["row_count"]}건) 피해 {row["damage_amount_central_confirmed"]:,}천원 / '
              f'복구비 {row["recovery_cost_central_confirmed"]:,}천원')
    for row in skipped:
        print(f'  [SKIP ] {row["record_id"]}: {row["reason"]}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
